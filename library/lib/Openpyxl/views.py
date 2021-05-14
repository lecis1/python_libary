from django.shortcuts import render, HttpResponse

from django.views.decorators.csrf import csrf_exempt
from django.http import FileResponse

import os
import openpyxl
from openpyxl.writer.excel import save_virtual_workbook
import io

from .models import User


# Create your views here.


@csrf_exempt
def import_data(request, *args, **kwargs):
    print(request.FILES.getlist('uploadFile'))
    files = request.FILES.getlist('uploadFile')
    try:
        file = files[0]
    except IndexError:
        return HttpResponse('读取文件失败！！')
        # return Response(data={'msg': '读取文件失败'},
        # status=status.HTTP_400_BAD_REQUEST)

    try:
        _, ext = os.path.splitext(file.name)
    except Exception as e:
        return HttpResponse('解析文件失败！！！')
        # return Response(data={'msg': '解析文件失败！'},
        # status=status.HTTP_400_BAD_REQUEST)

    if ext not in ['.xlsx', '.xls']:
        return HttpResponse('导入失败！！！！文件类型错误')
        # return Response(data={'msg': '不支持该类型文件的读取！'}, status=status.HTTP_400_BAD_REQUEST)
    wb = openpyxl.load_workbook(file)
    ws = wb.active
    objs = []
    # 重复数据
    repeated_data = []
    invalid_data = []

    for index, row_value in enumerate(ws.values):
        if index == 0:
            continue
        if len(row_value) != 5:
            invalid_data.append('第%d行, 数据的参数不正确' % (index + 1))
            continue

        name = row_value[0]
        count = row_value[1]
        sex = row_value[2]
        age = row_value[3]
        university = row_value[4]

        if not (name and count and sex and age and university):
            invalid_data.append('第%d行, 数据不能为空' % (index + 1))
            continue

        if User.objects.filter(count=count).exists():
            repeated_data.append('第%d行, 该条数据已存在' % (index + 1))

            continue

        data = {
            'name': name,
            'count': count,
            'sex': sex,
            'age': age,
            'university': university,
        }

        objs.append(User(**data))

    if any([repeated_data, invalid_data]):
        return HttpResponse("{'msg': '导入数据未通过','repeated: %s','invalid_data: "
                            "%s',}" % (repeated_data, invalid_data))

        # return Response(data={
        #     'msg': '导入数据未通过',
        #     'repeated': repeated_data,
        #     'invalid_data': invalid_data,
        # }, status=status.HTTP_400_BAD_REQUEST)

    User.objects.bulk_create(objs)
    return HttpResponse('导入成功')


@csrf_exempt
def export_data(request, *args, **kwargs):
    # 导出格式支持csv, excel,json.默认为excel
    fmt = request.GET.get('fmt', 'excel')
    default_file_name = '用户'
    filename = request.GET.get('filename', default_file_name)

    results = []
    queryset = User.objects.all()

    for record in queryset:
        results.append((
            record.name,
            record.count,
            record.sex,
            record.age,
            record.university,
        ))

    titles = ['用户名', '账号', '性别', '年龄', '学校']
    if fmt == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        for index, title in enumerate(titles):
            ws.cell(1, index+1, title)
        for record in results:
            ws.append(record)
        response = FileResponse(io.BytesIO(save_virtual_workbook(wb)), as_attachment=True,
                                filename=f'{filename}.xlsx')
        response['Access-Content-Expose-Headers'] = 'Content-Disposition'
        return response
    else:
        error_msg = f'不支持{fmt}导出格式'
        return HttpResponse("{'msg': %s}" % error_msg)