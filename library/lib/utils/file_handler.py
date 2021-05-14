import xlwt

import openpyxl


def write_data_to_wb(table_headers, files):
    wb = xlwt.Workbook(encoding='utf-8')
    ws = wb.add_sheet('sheet')
    ws.col(4).width = 8888
    # 建立样式
    font_style = xlwt.XFStyle()
    # 设置字体的样式
    font_style.font.bold = True
    # 设置为水平居中
    # horz(水平居中)、vert(垂直居中)、rote(旋转方向45°)、shri(自动缩进)
    font_style.alignment.horz = 2
    for header_index, header in enumerate(table_headers):
        ws.write(0, header_index, header, font_style)

    row_num = 0
    for file in files:
        row_num += 1
        for file_index, detail in enumerate(file):
            try:
                ws.write(row_num, file_index, detail, font_style,)
            except:
                ws.write(row_num, file_index, 'xxl', font_style)

    return wb


def readwb(file):
    wb = openpyxl.load_workbook(file)
    ws = wb.active

    objs = []
    invalid_data = []

    for index, row_value in enumerate(ws.values):
        if index == 0:
            continue
        if len(row_value) != 5:
            print('导入的数据的格式不正确,请检查！')
            invalid_data.append(index)


    for row in sheet.rows:
        row_data = [cell.value for cell in row]
        all_info.append(row_data)
    return sorted(all_info, key=lambda item: item[1])


def save_to_excel(data, wbname, sheetname='sheet1'):
    """
    将获取到的数据保存到excel中
    :param data:
    :param wbname:
    :param sheetname:
    :return:
    """

    print('导入excel[%s]中...' % wbname)
    wb = openpyxl.Workbook()
    sheet  = wb.active

    sheet.title = sheetname

    for row, item in enumerate(data):
        for column, cellValue in enumerate(item):
            sheet.cell(row=row+1, column=column+1, value=cellValue)

    wb.save(filename=wbname)

    print('写入成功！')


